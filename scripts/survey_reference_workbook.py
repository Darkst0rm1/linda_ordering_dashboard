"""Survey the reference workbook: sheet names, dims, formula counts, #REF!
counts, tab colors. Cheap first pass so we know where to dig deeper."""
import json
from pathlib import Path
import openpyxl

DATA = Path(__file__).resolve().parent.parent / "sample_data"
PATH = DATA / "Linda New Ordering Spreadsheet 08_19_2026.xlsx"

wb_formulas = openpyxl.load_workbook(PATH, read_only=True, data_only=False, keep_links=False)

report = {}
total_formulas = 0
total_ref_errors = 0

for name in wb_formulas.sheetnames:
    ws = wb_formulas[name]
    formula_count = 0
    ref_error_count = 0
    sample_ref_errors = []
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    # scan cells for formulas (read_only iter is memory friendly)
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                formula_count += 1
                if "#REF!" in v:
                    ref_error_count += 1
                    if len(sample_ref_errors) < 3:
                        sample_ref_errors.append({"cell": cell.coordinate, "formula": v[:200]})
    total_formulas += formula_count
    total_ref_errors += ref_error_count
    report[name] = {
        "max_row": max_row,
        "max_col": max_col,
        "formula_count": formula_count,
        "ref_error_count": ref_error_count,
        "sample_ref_errors": sample_ref_errors,
        "sheet_state": ws.sheet_state,
    }
    print(f"{name!r}: rows={max_row} cols={max_col} formulas={formula_count} refErrors={ref_error_count} state={ws.sheet_state}")

wb_formulas.close()

print("\nTOTAL formulas:", total_formulas)
print("TOTAL #REF! errors:", total_ref_errors)
print("Sheet count:", len(report))

out = DATA / "_reference_workbook_survey.json"
out.write_text(json.dumps(report, indent=2))
print("Wrote", out)
