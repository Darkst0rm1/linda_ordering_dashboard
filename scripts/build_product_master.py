"""Extract the product/vendor master from the 16 named supplier sheets of the
reference workbook into a version-controlled CSV config.

Rules (see Claude_Instructions_Streamlit_Linda_Ordering_Dashboard.md):
 - never translate a #REF! formula into dashboard logic
 - Material = value of the "SAP #" column when present; otherwise
   reconstructed as "100" + TOL # (this matches the workbook's own formula
   pattern, e.g. `=100&C4`, confirmed non-broken)
 - "Base Line" / "Current Month" / "Lost" columns are excluded entirely
   -- most are broken (#REF!) XLOOKUPs
 - Latest_Forecast_Qty = the right-most "Forecast AS400"-tagged column (by
   sheet column position) that has a non-null, non-error numeric value for
   that row. This is a legacy, non-guaranteed-chronological ordering -- it
   is documented as such, never silently re-sorted or "corrected".
"""
import csv
import json
import re
from pathlib import Path
import openpyxl

DATA = Path(__file__).resolve().parent.parent / "sample_data"
PATH = DATA / "Linda New Ordering Spreadsheet 08_19_2026.xlsx"
OUT = Path(__file__).resolve().parent.parent / "config" / "product_master.csv"
AUDIT_OUT = Path(__file__).resolve().parent.parent / "config" / "product_master_extraction_audit.json"

SUPPLIER_SHEETS = [
    "Alimentias", "Cheeseland", "Vergeer", "Woerle", "Freisland", "Caputo",
    "Bothwell", "Cows", "Anne of GG", "AppleKiss", "Sabana", "Stonetown",
    "La Tortilla", "Traditional", "Gay Lea", "Old Croc",
]

FORECAST_RE = re.compile(r"forecast|fcst", re.I)
ERROR_STRINGS = {"#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!"}


def clean_material(v):
    if v is None:
        return None
    s = str(v).strip()
    if s.endswith(".0") and s.replace(".0", "").isdigit():
        s = s[:-2]
    return s or None


def clean_tol(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    if s.endswith(".0") and s.replace(".0", "").isdigit():
        s = s[:-2]
    return s or None


wb_f = openpyxl.load_workbook(PATH, read_only=True, data_only=False, keep_links=False)
wb_v = openpyxl.load_workbook(PATH, read_only=True, data_only=True, keep_links=False)

rows_out = []
audit = {}

for name in SUPPLIER_SHEETS:
    ws_f = wb_f[name]
    ws_v = wb_v[name]
    header_row = 3
    max_col = ws_f.max_column
    max_row = ws_f.max_row

    plant_col = sap_col = tol_col = desc_col = None
    forecast_cols = []  # list of (col_idx, header_text)
    for c in range(1, max_col + 1):
        h = ws_f.cell(row=header_row, column=c).value
        if not isinstance(h, str):
            continue
        h_stripped = h.strip()
        # only take the FIRST occurrence of each label -- some sheets reuse
        # "SAP #" / "Product" as a header again far to the right for an
        # unrelated helper block, and that must not override the real
        # front-of-sheet product-identity columns
        if h_stripped == "Plant #" and plant_col is None:
            plant_col = c
        elif h_stripped == "SAP #" and sap_col is None:
            sap_col = c
        elif h_stripped == "TOL #" and tol_col is None:
            tol_col = c
        elif h_stripped in ("Product", "Item") and desc_col is None and tol_col is not None:
            # first Product/Item column after TOL # is the description
            desc_col = c
        if FORECAST_RE.search(h_stripped):
            forecast_cols.append((c, h_stripped))

    # a legitimate per-row "SAP #" identity column always sits immediately
    # next to Plant #/TOL #/Product at the front of the sheet; a "SAP #"
    # header re-used far to the right (e.g. next to a helper/lookup block)
    # is not a per-row material column, so fall back to TOL reconstruction
    if sap_col is not None and tol_col is not None and abs(sap_col - tol_col) > 3:
        sap_col = None

    reconstruct_material = sap_col is None

    n_rows_extracted = 0
    n_rows_skipped_blank = 0
    n_rows_missing_material = 0
    forecast_source_counts = {}

    for r in range(header_row + 1, max_row + 1):
        plant = ws_v.cell(row=r, column=plant_col).value if plant_col else None
        tol_raw = ws_v.cell(row=r, column=tol_col).value if tol_col else None
        desc = ws_v.cell(row=r, column=desc_col).value if desc_col else None

        if plant is None and tol_raw is None and desc is None:
            n_rows_skipped_blank += 1
            continue

        if plant is None:
            # A populated TOL#/Product/forecast with no Plant value is not a
            # real per-row product -- it's bleed from a differently-laid-out
            # block elsewhere in the sheet (e.g. a legend/notes area). The
            # (Plant, Material) composite key requires a real plant, so these
            # rows are skipped rather than kept with a blank/invented plant.
            n_rows_skipped_blank += 1
            continue

        tol = clean_tol(tol_raw)

        if reconstruct_material:
            material = f"100{tol}" if tol else None
        else:
            material = clean_material(ws_v.cell(row=r, column=sap_col).value)

        if material is None:
            n_rows_missing_material += 1
            continue

        latest_val = None
        latest_header = None
        for c, h in reversed(forecast_cols):
            v = ws_v.cell(row=r, column=c).value
            if v is None:
                continue
            if isinstance(v, str):
                if v.strip() in ERROR_STRINGS or not v.strip():
                    continue
                try:
                    v = float(v)
                except ValueError:
                    continue
            if isinstance(v, (int, float)):
                latest_val = v
                latest_header = h
                break

        if latest_header:
            forecast_source_counts[latest_header] = forecast_source_counts.get(latest_header, 0) + 1

        rows_out.append({
            "Supplier": name,
            "Plant": clean_material(plant),
            "Material": material,
            "TOL_Material": tol,
            "Description": (desc.strip() if isinstance(desc, str) else desc) or "",
            "Latest_Forecast_Qty": latest_val if latest_val is not None else "",
            "Forecast_Source_Column": latest_header or "",
        })
        n_rows_extracted += 1

    audit[name] = {
        "header_row": header_row,
        "plant_col": plant_col,
        "sap_col": sap_col,
        "tol_col": tol_col,
        "desc_col": desc_col,
        "reconstructed_material_from_tol": reconstruct_material,
        "forecast_col_count": len(forecast_cols),
        "rows_extracted": n_rows_extracted,
        "rows_skipped_blank": n_rows_skipped_blank,
        "rows_skipped_missing_material": n_rows_missing_material,
        "forecast_source_column_distribution": forecast_source_counts,
    }
    print(f"{name}: extracted={n_rows_extracted} blank_skipped={n_rows_skipped_blank} "
          f"missing_material={n_rows_missing_material} sap_col={sap_col} tol_col={tol_col} desc_col={desc_col}")

wb_f.close()
wb_v.close()

OUT.parent.mkdir(exist_ok=True)
with OUT.open("w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=[
        "Supplier", "Plant", "Material", "TOL_Material", "Description",
        "Latest_Forecast_Qty", "Forecast_Source_Column",
    ])
    writer.writeheader()
    writer.writerows(rows_out)

audit["_known_limitations"] = {
    "Caputo": (
        "Caputo (0 rows extracted) uses a block/label layout, not the uniform "
        "one-row-per-material table used by the other 15 supplier sheets: the "
        "TOL #/Product columns are blank for essentially every data row, and "
        "the real product identity (e.g. \"71000071 Wiscon Corporation K20\") "
        "appears as a scattered label row above each block instead. No merged "
        "cells tie the label to its block (verified via openpyxl merged_cells), "
        "so there is no safe, non-invented way to propagate it down to the "
        "per-row per-plant forecast entries. Materials sourced only from "
        "Caputo therefore have Target Stock = blank (\"Needs planner target\") "
        "until a planner manually maps this sheet. Cows uses the same front "
        "columns (Plant #, TOL #, Product) as the uniform sheets and DID "
        "extract 39 rows successfully, so it is not affected by this limitation."
    ),
    "Cheeseland_and_La_Tortilla": (
        "15 rows originally extracted from these two sheets had a populated "
        "TOL #/Product but a blank Plant # -- bleed from a secondary "
        "legend/notes block elsewhere in the sheet with a different column "
        "layout, not real per-plant product rows. They are skipped entirely "
        "(counted in rows_skipped_blank above) rather than kept with a "
        "blank/invented Plant, since (Plant, Material) is a required "
        "composite key everywhere downstream."
    ),
}

AUDIT_OUT.write_text(json.dumps(audit, indent=2))
print(f"\nWrote {len(rows_out)} product master rows to {OUT}")
print(f"Wrote extraction audit to {AUDIT_OUT}")
