"""One-off inspection script: validate the 9 supplied SAP exports against the
required schemas from the instructions doc. Not part of the app itself."""
import json
from pathlib import Path
import openpyxl

DATA = Path(__file__).resolve().parent.parent / "sample_data"

CUSTOMER_ORDERS_COLS = [
    "Plant", "Sales Order", "Sales Order Type Desc.", "Sold To Name", "Material",
    "TOL Material Description", "Order Quantity", "Confirmed Quantity (CS)",
    "Reason for Rejection Desc.", "Picked Quantity (CS)", "Picked Quantity (KG)",
    "Invoice Quantity (CS)", "Invoice Quantity (KG)", "Item Net Amount",
    "Outbound Delivery #", "Outbound Delivery Status", "BBD / Shelf Life",
    "Item Net Amount (Confirmed)", "Requested Delivery Date", "Ship to Arrive Date",
    "Appointment Confirmation", "Shipped Date", "Vendor", "Vendor Name",
    "Material Group Description", "Purchasing Group", "Purchasing Group Name",
    "BDM Description", "CDM Name", "Credit Check Status", "Credit Check Status Desc.",
]

ON_HAND_COLS = [
    "Material", "Material Description", "Plant", "Plant Name", "Storage Location",
    "Stock in Quality Inspection", "Unrestricted Stock", "Batch", "Blocked Stock",
    "Production Date", "Shelf Life Expiration Date",
]

OPEN_ORDERS_COLS = [
    "Material", "PO Number", "Open PO Qty", "PO Quantity", "PO Request Quantity",
    "PO Received Qty", "Inbound Delivery Quantity", "Company Code", "Vendor",
    "Vendor Name", "Supplier", "Supplier Name", "Delivery Date", "Est PU Date",
    "Plant", "SLOC", "Material Description", "Material Group", "Delivery Completed",
    "Created On Date", "Appt. Plant", "Gross Weight", "Appt. Date", "Appt. Time",
    "Delivery Date Derived Field", "Delivery Priority Text", "Inbound Delivery Status",
    "STO/CD Pro #",
]

FILES = {
    "2910Customerorders.xlsx": ("customer_orders", "2910", CUSTOMER_ORDERS_COLS),
    "2920CustomerOrders.xlsx": ("customer_orders", "2920", CUSTOMER_ORDERS_COLS),
    "2930Customerorders.xlsx": ("customer_orders", "2930", CUSTOMER_ORDERS_COLS),
    "OH2910.xlsx": ("on_hand", "2910", ON_HAND_COLS),
    "OH20.xlsx": ("on_hand", "2920", ON_HAND_COLS),
    "OH2930.xlsx": ("on_hand", "2930", ON_HAND_COLS),
    "OR2910.xlsx": ("open_orders", "2910", OPEN_ORDERS_COLS),
    "OR2920.xlsx": ("open_orders", "2920", OPEN_ORDERS_COLS),
    "OR2930.xlsx": ("open_orders", "2930", OPEN_ORDERS_COLS),
}

report = {}
for fname, (kind, plant, expected_cols) in FILES.items():
    path = DATA / fname
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    has_sheet = "SAPUI5 Export" in sheet_names
    ws = wb["SAPUI5 Export"] if has_sheet else wb[sheet_names[0]]
    rows = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header = list(rows[0]) if rows else []
    header = [h for h in header]
    # trim trailing Nones
    while header and header[-1] is None:
        header.pop()
    row_count = ws.max_row
    plant_values = set()
    plant_col_idx = None
    if header:
        for i, h in enumerate(header):
            if h == "Plant":
                plant_col_idx = i
                break
    if plant_col_idx is not None:
        for r in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 2000), values_only=True):
            if plant_col_idx < len(r):
                plant_values.add(r[plant_col_idx])
    report[fname] = {
        "kind": kind,
        "expected_plant": plant,
        "sheet_names": sheet_names,
        "has_SAPUI5_Export_sheet": has_sheet,
        "header_matches_exactly": header == expected_cols,
        "header_len": len(header),
        "expected_len": len(expected_cols),
        "header_actual": header,
        "missing_cols": [c for c in expected_cols if c not in header],
        "extra_cols": [c for c in header if c not in expected_cols],
        "row_count_incl_header": row_count,
        "data_row_count": max(row_count - 1, 0),
        "distinct_plant_values_sampled": sorted([str(v) for v in plant_values if v is not None]),
    }
    wb.close()

out_path = Path(__file__).resolve().parent.parent / "sample_data" / "_sap_export_inspection.json"
out_path.write_text(json.dumps(report, indent=2, default=str))
print(json.dumps(report, indent=2, default=str))
