"""Deep-dive extraction for the 16 named supplier sheets plus the forecast /
PIR sheets that might hold target-stock baselines. Writes one JSON per sheet
under sample_data/_extract/ so we can inspect structure without blowing up
the conversation context with raw cell dumps.

Strategy per sheet:
 - scan the first 6 rows fully (formulas) to find the real header row (row
   with the most non-null string cells that look like labels)
 - record header row values + column letter + fill color + font color
 - for the header row's columns, sample the first 15 data rows' *values*
   (data_only workbook) so #REF! formulas resolve to their last cached
   value/error string instead of the formula text
 - record tab color
 - flag columns whose header text matches target/safety/forecast/case-pack
   keywords for later business-rule extraction
"""
import json
import re
from pathlib import Path
import openpyxl

DATA = Path(__file__).resolve().parent.parent / "sample_data"
PATH = DATA / "Linda New Ordering Spreadsheet 08_19_2026.xlsx"
OUT_DIR = DATA / "_extract"
OUT_DIR.mkdir(exist_ok=True)

SUPPLIER_SHEETS = [
    "Alimentias", "Cheeseland", "Vergeer", "Woerle", "Freisland", "Caputo",
    "Bothwell", "Cows", "Anne of GG", "AppleKiss", "Sabana", "Stonetown",
    "La Tortilla", "Traditional", "Gay Lea", "Old Croc",
]

KEYWORD_PATTERNS = {
    "material": re.compile(r"material|sap\s*#|sap\s*material", re.I),
    "tol": re.compile(r"\btol\b", re.I),
    "description": re.compile(r"description|desc\.?$", re.I),
    "vendor": re.compile(r"vendor|supplier", re.I),
    "plant": re.compile(r"plant", re.I),
    "forecast": re.compile(r"forecast|fcst", re.I),
    "target": re.compile(r"target|par\s*level|min.?max|reorder", re.I),
    "safety": re.compile(r"safety", re.I),
    "case_pack": re.compile(r"case\s*pack|units?/case|per\s*case", re.I),
    "on_hand": re.compile(r"on\s*hand|unrestricted", re.I),
    "open_po": re.compile(r"open\s*po|on\s*order", re.I),
}

wb_f = openpyxl.load_workbook(PATH, read_only=False, data_only=False)
wb_v = openpyxl.load_workbook(PATH, read_only=False, data_only=True)

def classify(header_text):
    if not header_text:
        return []
    tags = []
    for tag, pat in KEYWORD_PATTERNS.items():
        if pat.search(str(header_text)):
            tags.append(tag)
    return tags

def find_header_row(ws_f, max_scan_rows=8, max_scan_cols=80):
    best_row, best_score = 1, -1
    for r in range(1, max_scan_rows + 1):
        score = 0
        for c in range(1, max_scan_cols + 1):
            v = ws_f.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip() and not v.startswith("="):
                score += 1
        if score > best_score:
            best_score = score
            best_row = r
    return best_row

for name in SUPPLIER_SHEETS:
    ws_f = wb_f[name]
    ws_v = wb_v[name]
    tab_color = None
    try:
        tc = ws_f.sheet_properties.tabColor
        tab_color = tc.rgb if tc else None
    except Exception:
        pass

    header_row = find_header_row(ws_f)
    max_col = min(ws_f.max_column, 120)
    max_row = ws_f.max_row

    columns = []
    for c in range(1, max_col + 1):
        cell = ws_f.cell(row=header_row, column=c)
        v = cell.value
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        fill = None
        try:
            fg = cell.fill.fgColor
            fill = fg.rgb if fg and fg.type == "rgb" else (str(fg.theme) if fg and fg.type == "theme" else None)
        except Exception:
            pass
        font_color = None
        try:
            fc = cell.font.color
            font_color = fc.rgb if fc and fc.type == "rgb" else None
        except Exception:
            pass
        columns.append({
            "col_idx": c,
            "col_letter": cell.coordinate[: len(cell.coordinate) - len(str(header_row))],
            "header": v,
            "tags": classify(v),
            "fill": fill,
            "font_color": font_color,
        })

    # sample data rows (values, not formulas) for the tagged columns only,
    # to keep payload small
    tagged_cols = [c for c in columns if c["tags"]]
    sample_rows = []
    max_sample = min(max_row, header_row + 15)
    for r in range(header_row + 1, max_sample + 1):
        row_out = {}
        any_val = False
        for c in tagged_cols:
            val = ws_v.cell(row=r, column=c["col_idx"]).value
            row_out[c["header"]] = val
            if val is not None:
                any_val = True
        if any_val:
            sample_rows.append(row_out)

    result = {
        "sheet": name,
        "tab_color": tab_color,
        "max_row": max_row,
        "max_col_scanned": max_col,
        "header_row": header_row,
        "columns_all": columns,
        "columns_tagged": tagged_cols,
        "sample_data_rows": sample_rows,
    }
    out_path = OUT_DIR / f"{name.replace(' ', '_').replace('/', '_')}.json"
    out_path.write_text(json.dumps(result, indent=2, default=str))
    print(f"{name}: header_row={header_row} tagged_cols={[c['header'] for c in tagged_cols]}")

wb_f.close()
wb_v.close()
print("Done. Wrote per-sheet JSON to", OUT_DIR)
