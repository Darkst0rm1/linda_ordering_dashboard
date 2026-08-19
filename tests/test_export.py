import io
import zipfile

import openpyxl
import pandas as pd

from src.aggregate import build_ordering_dashboard
from src.excel_export import build_export_workbook, build_audit_report_text
from src.reconciliation import build_reconciliation_report


def _build(on_hand_df, customer_orders_df, open_orders_df, target_stock_df, allocation_field_by_plant):
    grid = build_ordering_dashboard(
        on_hand_df, customer_orders_df, open_orders_df, target_stock_df,
        allocation_field_by_plant, safety_stock_pct=0.20,
    )
    raw_frames = {"on_hand": on_hand_df, "customer_orders": customer_orders_df, "open_orders": open_orders_df}
    report = build_reconciliation_report(raw_frames, {}, grid, target_stock_df)
    return grid, raw_frames, report


def test_export_workbook_opens_without_error(
    on_hand_df, customer_orders_df, open_orders_df, target_stock_df, allocation_field_by_plant,
):
    grid, raw_frames, report = _build(on_hand_df, customer_orders_df, open_orders_df, target_stock_df, allocation_field_by_plant)
    xlsx_bytes = build_export_workbook(grid, raw_frames, report, "2026-08-19 12:00:00 EDT")

    # programmatic read-back -- openpyxl raises on a corrupt/needs-repair file
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert "Executive Summary" in wb.sheetnames
    assert "Ordering Dashboard" in wb.sheetnames
    assert "Reconciliation" in wb.sheetnames
    assert "Unmatched Materials" in wb.sheetnames
    assert "Data Dictionary" in wb.sheetnames
    for kind_label in ["Customer Orders", "On Hand", "Open Orders"]:
        for plant in ["2910", "2920", "2930"]:
            assert f"{kind_label} {plant}" in wb.sheetnames


def test_export_workbook_has_supplier_sheets(
    on_hand_df, customer_orders_df, open_orders_df, target_stock_df, allocation_field_by_plant,
):
    grid, raw_frames, report = _build(on_hand_df, customer_orders_df, open_orders_df, target_stock_df, allocation_field_by_plant)
    xlsx_bytes = build_export_workbook(grid, raw_frames, report, "2026-08-19 12:00:00 EDT")
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert "Alimentias" in wb.sheetnames


def test_export_workbook_has_no_external_links(
    on_hand_df, customer_orders_df, open_orders_df, target_stock_df, allocation_field_by_plant,
):
    grid, raw_frames, report = _build(on_hand_df, customer_orders_df, open_orders_df, target_stock_df, allocation_field_by_plant)
    xlsx_bytes = build_export_workbook(grid, raw_frames, report, "2026-08-19 12:00:00 EDT")
    with zipfile.ZipFile(io.BytesIO(xlsx_bytes)) as z:
        assert not any("externalLink" in name for name in z.namelist())


def test_export_workbook_has_no_pivot_tables(
    on_hand_df, customer_orders_df, open_orders_df, target_stock_df, allocation_field_by_plant,
):
    grid, raw_frames, report = _build(on_hand_df, customer_orders_df, open_orders_df, target_stock_df, allocation_field_by_plant)
    xlsx_bytes = build_export_workbook(grid, raw_frames, report, "2026-08-19 12:00:00 EDT")
    with zipfile.ZipFile(io.BytesIO(xlsx_bytes)) as z:
        assert not any("pivotTable" in name for name in z.namelist())


def test_audit_report_text_contains_reconciliation_status(
    on_hand_df, customer_orders_df, open_orders_df, target_stock_df, allocation_field_by_plant,
):
    grid, raw_frames, report = _build(on_hand_df, customer_orders_df, open_orders_df, target_stock_df, allocation_field_by_plant)
    text = build_audit_report_text(report, "2026-08-19 12:00:00 EDT")
    assert "AUDIT REPORT" in text
    assert "PASS" in text or "FAIL" in text
